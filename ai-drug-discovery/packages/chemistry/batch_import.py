"""
Batch CSV/Excel ingestion with column mapping.

This module provides robust batch import of molecular data from tabular files:
- Supports CSV and Excel (.xlsx, .xls) files
- Flexible column mapping with auto-detection
- Row-level validation with detailed error reporting
- Integrates with SMILES validation pipeline

Column Mapping:
---------------
Provide a mapping dict to specify which columns contain molecular data:

    mapping = {
        "smiles_col": "SMILES",      # Required: column with SMILES strings
        "name_col": "Compound_Name",  # Optional: molecule name
        "id_col": "External_ID",      # Optional: external identifier
    }

Auto-detection will try common column names if mapping is not provided.

Usage:
    >>> from packages.chemistry.batch_import import import_molecules_from_file

    # With explicit mapping
    >>> result = import_molecules_from_file(
    ...     "compounds.csv",
    ...     mapping={"smiles_col": "SMILES", "name_col": "Name"}
    ... )
    >>> print(f"Imported {result.success_count} molecules")

    # With auto-detection
    >>> result = import_molecules_from_file("compounds.xlsx")
    >>> for mol in result.molecules:
    ...     print(f"{mol.name}: {mol.smiles}")

    # Check errors
    >>> for err in result.errors:
    ...     print(f"Row {err.row_number}: {err.message}")
"""

import csv
import io
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any, BinaryIO, TextIO

# Lazy imports for optional dependencies
_pandas_available: bool | None = None
_openpyxl_available: bool | None = None


def _check_pandas() -> bool:
    """Check if pandas is available."""
    global _pandas_available
    if _pandas_available is None:
        try:
            import pandas  # noqa: F401

            _pandas_available = True
        except ImportError:
            _pandas_available = False
    return _pandas_available


def _check_openpyxl() -> bool:
    """Check if openpyxl is available."""
    global _openpyxl_available
    if _openpyxl_available is None:
        try:
            import openpyxl  # noqa: F401

            _openpyxl_available = True
        except ImportError:
            _openpyxl_available = False
    return _openpyxl_available


# =============================================================================
# Constants
# =============================================================================

# Common column names for auto-detection (case-insensitive)
SMILES_COLUMN_NAMES = [
    "smiles",
    "canonical_smiles",
    "isomeric_smiles",
    "structure",
    "molecule",
    "mol",
    "compound",
    "smi",
    "input_smiles",
]

NAME_COLUMN_NAMES = [
    "name",
    "compound_name",
    "molecule_name",
    "mol_name",
    "title",
    "common_name",
    "preferred_name",
    "generic_name",
]

ID_COLUMN_NAMES = [
    "id",
    "compound_id",
    "molecule_id",
    "mol_id",
    "external_id",
    "ext_id",
    "cas",
    "cas_number",
    "chembl_id",
    "pubchem_cid",
    "registry_number",
]


class FileType(str, Enum):
    """Supported file types."""

    CSV = "csv"
    TSV = "tsv"
    EXCEL = "excel"
    UNKNOWN = "unknown"


class ImportErrorCode(str, Enum):
    """Error codes for batch import."""

    # File-level errors
    FILE_NOT_FOUND = "FILE_NOT_FOUND"
    FILE_READ_ERROR = "FILE_READ_ERROR"
    UNSUPPORTED_FORMAT = "UNSUPPORTED_FORMAT"
    EMPTY_FILE = "EMPTY_FILE"

    # Mapping errors
    MISSING_SMILES_COLUMN = "MISSING_SMILES_COLUMN"
    INVALID_COLUMN_MAPPING = "INVALID_COLUMN_MAPPING"

    # Row-level errors
    EMPTY_SMILES = "EMPTY_SMILES"
    INVALID_SMILES = "INVALID_SMILES"
    VALIDATION_FAILED = "VALIDATION_FAILED"


# =============================================================================
# Data Classes
# =============================================================================


@dataclass
class ColumnMapping:
    """
    Column mapping configuration.

    Specifies which columns contain molecular data.
    """

    smiles_col: str  # Required
    name_col: str | None = None
    id_col: str | None = None

    # Additional custom columns to extract
    extra_cols: dict[str, str] = field(default_factory=dict)

    def validate(self, available_columns: list[str]) -> list[str]:
        """
        Validate mapping against available columns.

        Args:
            available_columns: List of column names in the file.

        Returns:
            List of error messages (empty if valid).
        """
        errors = []
        cols_lower = {c.lower(): c for c in available_columns}

        # Check required SMILES column
        if self.smiles_col.lower() not in cols_lower:
            errors.append(
                f"SMILES column '{self.smiles_col}' not found. "
                f"Available: {available_columns}"
            )

        # Check optional columns
        if self.name_col and self.name_col.lower() not in cols_lower:
            errors.append(f"Name column '{self.name_col}' not found")

        if self.id_col and self.id_col.lower() not in cols_lower:
            errors.append(f"ID column '{self.id_col}' not found")

        # Check extra columns
        for key, col in self.extra_cols.items():
            if col.lower() not in cols_lower:
                errors.append(f"Extra column '{col}' (mapped as '{key}') not found")

        return errors

    def get_actual_column_name(
        self, col_name: str, available_columns: list[str]
    ) -> str | None:
        """Get the actual column name (case-matched) from available columns."""
        cols_lower = {c.lower(): c for c in available_columns}
        return cols_lower.get(col_name.lower())


@dataclass
class ImportedMolecule:
    """
    Successfully imported molecule record.

    Contains the validated SMILES and extracted metadata.
    """

    row_number: int  # 1-based row number (header is row 1)
    smiles: str  # Validated SMILES (may be original or canonicalized)
    name: str | None = None
    external_id: str | None = None
    extra_data: dict[str, Any] = field(default_factory=dict)

    # Validation results (if canonicalization was performed)
    canonical_smiles: str | None = None
    inchikey: str | None = None
    is_valid: bool = True
    warnings: list[str] = field(default_factory=list)


@dataclass
class RowError:
    """
    Error record for a failed row.

    Contains row location and detailed error information.
    """

    row_number: int  # 1-based row number
    error_code: ImportErrorCode
    message: str
    smiles_value: str | None = None  # The problematic SMILES (if available)
    column_name: str | None = None  # Column that caused the error
    details: dict[str, Any] = field(default_factory=dict)


@dataclass
class BatchImportResult:
    """
    Result of batch import operation.

    Contains successfully imported molecules and row-level errors.
    """

    # Successfully imported molecules
    molecules: list[ImportedMolecule] = field(default_factory=list)

    # Row-level errors
    errors: list[RowError] = field(default_factory=list)

    # Statistics
    total_rows: int = 0
    success_count: int = 0
    error_count: int = 0

    # File info
    file_path: str | None = None
    file_type: FileType = FileType.UNKNOWN
    detected_columns: list[str] = field(default_factory=list)
    used_mapping: ColumnMapping | None = None

    @property
    def success_rate(self) -> float:
        """Calculate success rate as percentage."""
        if self.total_rows == 0:
            return 0.0
        return (self.success_count / self.total_rows) * 100.0

    @property
    def has_errors(self) -> bool:
        """Check if any errors occurred."""
        return self.error_count > 0


@dataclass
class FileReadResult:
    """Result of reading a tabular file."""

    headers: list[str]
    rows: list[dict[str, Any]]
    file_type: FileType
    total_rows: int
    error: str | None = None


# =============================================================================
# File Detection and Reading
# =============================================================================


def detect_file_type(
    filepath: str | Path | None = None,
    filename: str | None = None,
    content: bytes | None = None,
) -> FileType:
    """
    Detect file type from path, filename, or content.

    Args:
        filepath: File path (optional).
        filename: Filename for extension detection (optional).
        content: File content for magic byte detection (optional).

    Returns:
        Detected FileType.
    """
    # Try extension-based detection
    name = None
    if filepath:
        name = str(filepath).lower()
    elif filename:
        name = filename.lower()

    if name:
        if name.endswith(".csv"):
            return FileType.CSV
        elif name.endswith(".tsv") or name.endswith(".txt"):
            return FileType.TSV
        elif name.endswith((".xlsx", ".xls", ".xlsm")):
            return FileType.EXCEL

    # Try content-based detection
    if content:
        # Excel files start with PK (ZIP) or specific bytes
        if content[:4] == b"PK\x03\x04":  # ZIP (xlsx)
            return FileType.EXCEL
        if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # OLE (xls)
            return FileType.EXCEL
        # Assume CSV/TSV for text content
        try:
            text = content[:1000].decode("utf-8")
            if "\t" in text and "," not in text:
                return FileType.TSV
            return FileType.CSV
        except UnicodeDecodeError:
            pass

    return FileType.UNKNOWN


def read_csv_file(
    source: str | Path | TextIO | BinaryIO,
    delimiter: str = ",",
    encoding: str = "utf-8",
) -> FileReadResult:
    """
    Read CSV/TSV file into list of dicts.

    Args:
        source: File path or file-like object.
        delimiter: Column delimiter.
        encoding: Text encoding.

    Returns:
        FileReadResult with headers and rows.
    """
    try:
        # Handle different source types
        if isinstance(source, (str, Path)):
            with open(source, "r", encoding=encoding, newline="") as f:
                return _read_csv_content(f, delimiter)
        elif hasattr(source, "read"):
            # File-like object
            content = source.read()
            if isinstance(content, bytes):
                content = content.decode(encoding)
            return _read_csv_content(io.StringIO(content), delimiter)
        else:
            return FileReadResult(
                headers=[],
                rows=[],
                file_type=FileType.CSV,
                total_rows=0,
                error=f"Unsupported source type: {type(source)}",
            )
    except Exception as e:
        return FileReadResult(
            headers=[],
            rows=[],
            file_type=FileType.CSV,
            total_rows=0,
            error=f"Failed to read CSV: {e}",
        )


def _read_csv_content(file_obj: TextIO, delimiter: str) -> FileReadResult:
    """Read CSV content from file object."""
    reader = csv.DictReader(file_obj, delimiter=delimiter)
    headers = reader.fieldnames or []
    rows = list(reader)

    return FileReadResult(
        headers=headers,
        rows=rows,
        file_type=FileType.CSV if delimiter == "," else FileType.TSV,
        total_rows=len(rows),
    )


def read_excel_file(
    source: str | Path | BinaryIO | bytes,
    sheet_name: str | int = 0,
) -> FileReadResult:
    """
    Read Excel file into list of dicts.

    Uses openpyxl for .xlsx files.

    Args:
        source: File path, file-like object, or bytes.
        sheet_name: Sheet name or index (0-based).

    Returns:
        FileReadResult with headers and rows.
    """
    if not _check_openpyxl():
        return FileReadResult(
            headers=[],
            rows=[],
            file_type=FileType.EXCEL,
            total_rows=0,
            error="openpyxl is required for Excel files. Install with: pip install openpyxl",
        )

    import openpyxl

    try:
        # Load workbook
        if isinstance(source, bytes):
            wb = openpyxl.load_workbook(
                io.BytesIO(source), read_only=True, data_only=True
            )
        elif isinstance(source, (str, Path)):
            wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        else:
            # File-like object
            wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

        # Get sheet
        if isinstance(sheet_name, int):
            if sheet_name >= len(wb.sheetnames):
                return FileReadResult(
                    headers=[],
                    rows=[],
                    file_type=FileType.EXCEL,
                    total_rows=0,
                    error=f"Sheet index {sheet_name} out of range",
                )
            ws = wb[wb.sheetnames[sheet_name]]
        else:
            if sheet_name not in wb.sheetnames:
                return FileReadResult(
                    headers=[],
                    rows=[],
                    file_type=FileType.EXCEL,
                    total_rows=0,
                    error=f"Sheet '{sheet_name}' not found",
                )
            ws = wb[sheet_name]

        # Read rows
        rows_iter = ws.iter_rows(values_only=True)

        # Get headers
        try:
            header_row = next(rows_iter)
            headers = [str(h).strip() if h is not None else f"col_{i}"
                      for i, h in enumerate(header_row)]
        except StopIteration:
            return FileReadResult(
                headers=[],
                rows=[],
                file_type=FileType.EXCEL,
                total_rows=0,
                error="Excel sheet is empty",
            )

        # Read data rows
        rows = []
        for row in rows_iter:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    # Convert to string, handle None
                    row_dict[headers[i]] = str(value).strip() if value is not None else ""
            rows.append(row_dict)

        wb.close()

        return FileReadResult(
            headers=headers,
            rows=rows,
            file_type=FileType.EXCEL,
            total_rows=len(rows),
        )

    except Exception as e:
        return FileReadResult(
            headers=[],
            rows=[],
            file_type=FileType.EXCEL,
            total_rows=0,
            error=f"Failed to read Excel file: {e}",
        )


def read_tabular_file(
    source: str | Path | TextIO | BinaryIO | bytes,
    file_type: FileType | None = None,
    delimiter: str | None = None,
    sheet_name: str | int = 0,
) -> FileReadResult:
    """
    Read tabular file (CSV, TSV, or Excel) with automatic type detection.

    Args:
        source: File path, file-like object, or bytes.
        file_type: Explicit file type (auto-detect if None).
        delimiter: CSV delimiter (auto-detect if None).
        sheet_name: Excel sheet name or index.

    Returns:
        FileReadResult with headers and rows.
    """
    # Detect file type if not specified
    if file_type is None:
        if isinstance(source, (str, Path)):
            file_type = detect_file_type(filepath=source)
        elif isinstance(source, bytes):
            file_type = detect_file_type(content=source)
        else:
            # Try to get filename from file-like object
            name = getattr(source, "name", None)
            file_type = detect_file_type(filename=name) if name else FileType.CSV

    # Read based on type
    if file_type == FileType.EXCEL:
        return read_excel_file(source, sheet_name=sheet_name)
    elif file_type in (FileType.CSV, FileType.TSV, FileType.UNKNOWN):
        delim = delimiter or ("," if file_type != FileType.TSV else "\t")
        return read_csv_file(source, delimiter=delim)
    else:
        return FileReadResult(
            headers=[],
            rows=[],
            file_type=file_type,
            total_rows=0,
            error=f"Unsupported file type: {file_type}",
        )


# =============================================================================
# Column Mapping
# =============================================================================


def auto_detect_mapping(columns: list[str]) -> ColumnMapping | None:
    """
    Auto-detect column mapping from column names.

    Args:
        columns: List of column names from file.

    Returns:
        ColumnMapping if SMILES column found, None otherwise.
    """
    cols_lower = {c.lower(): c for c in columns}

    # Find SMILES column (required)
    smiles_col = None
    for candidate in SMILES_COLUMN_NAMES:
        if candidate in cols_lower:
            smiles_col = cols_lower[candidate]
            break

    if not smiles_col:
        return None

    # Find name column (optional)
    name_col = None
    for candidate in NAME_COLUMN_NAMES:
        if candidate in cols_lower:
            name_col = cols_lower[candidate]
            break

    # Find ID column (optional)
    id_col = None
    for candidate in ID_COLUMN_NAMES:
        if candidate in cols_lower:
            id_col = cols_lower[candidate]
            break

    return ColumnMapping(
        smiles_col=smiles_col,
        name_col=name_col,
        id_col=id_col,
    )


def create_mapping(
    mapping_dict: dict[str, str],
    available_columns: list[str],
) -> tuple[ColumnMapping | None, list[str]]:
    """
    Create ColumnMapping from dict and validate.

    Args:
        mapping_dict: Mapping dict with keys like "smiles_col", "name_col", "id_col".
        available_columns: Available columns in the file.

    Returns:
        Tuple of (ColumnMapping or None, list of error messages).
    """
    errors = []

    # Extract standard mappings
    smiles_col = mapping_dict.get("smiles_col") or mapping_dict.get("smiles")
    name_col = mapping_dict.get("name_col") or mapping_dict.get("name")
    id_col = mapping_dict.get("id_col") or mapping_dict.get("id") or mapping_dict.get("external_id")

    if not smiles_col:
        errors.append("mapping_dict must contain 'smiles_col' key")
        return None, errors

    # Extract extra columns
    extra_cols = {}
    reserved_keys = {"smiles_col", "smiles", "name_col", "name", "id_col", "id", "external_id"}
    for key, col in mapping_dict.items():
        if key not in reserved_keys:
            extra_cols[key] = col

    mapping = ColumnMapping(
        smiles_col=smiles_col,
        name_col=name_col,
        id_col=id_col,
        extra_cols=extra_cols,
    )

    # Validate
    validation_errors = mapping.validate(available_columns)
    errors.extend(validation_errors)

    if errors:
        return None, errors

    return mapping, []


# =============================================================================
# Batch Import
# =============================================================================


class BatchImporter:
    """
    Batch importer for molecular data from tabular files.

    Handles CSV, TSV, and Excel files with configurable column mapping.
    """

    def __init__(
        self,
        validate_smiles: bool = True,
        canonicalize: bool = False,
        strip_whitespace: bool = True,
        skip_empty_rows: bool = True,
    ):
        """
        Initialize batch importer.

        Args:
            validate_smiles: Whether to validate SMILES strings.
            canonicalize: Whether to canonicalize SMILES (slower but provides InChIKey).
            strip_whitespace: Whether to strip whitespace from values.
            skip_empty_rows: Whether to skip rows with empty SMILES.
        """
        self.validate_smiles = validate_smiles
        self.canonicalize = canonicalize
        self.strip_whitespace = strip_whitespace
        self.skip_empty_rows = skip_empty_rows

    def import_from_file(
        self,
        source: str | Path | TextIO | BinaryIO | bytes,
        mapping: ColumnMapping | dict[str, str] | None = None,
        file_type: FileType | None = None,
        sheet_name: str | int = 0,
    ) -> BatchImportResult:
        """
        Import molecules from a tabular file.

        Args:
            source: File path, file-like object, or bytes.
            mapping: Column mapping (auto-detect if None).
            file_type: File type (auto-detect if None).
            sheet_name: Excel sheet name or index.

        Returns:
            BatchImportResult with molecules and errors.
        """
        # Read file
        file_result = read_tabular_file(
            source,
            file_type=file_type,
            sheet_name=sheet_name,
        )

        if file_result.error:
            return BatchImportResult(
                errors=[
                    RowError(
                        row_number=0,
                        error_code=ImportErrorCode.FILE_READ_ERROR,
                        message=file_result.error,
                    )
                ],
                error_count=1,
                file_type=file_result.file_type,
                file_path=str(source) if isinstance(source, (str, Path)) else None,
            )

        if not file_result.headers:
            return BatchImportResult(
                errors=[
                    RowError(
                        row_number=0,
                        error_code=ImportErrorCode.EMPTY_FILE,
                        message="File has no headers or is empty",
                    )
                ],
                error_count=1,
                file_type=file_result.file_type,
                detected_columns=[],
            )

        # Resolve mapping
        col_mapping: ColumnMapping | None = None
        mapping_errors: list[str] = []

        if mapping is None:
            # Auto-detect
            col_mapping = auto_detect_mapping(file_result.headers)
            if col_mapping is None:
                return BatchImportResult(
                    errors=[
                        RowError(
                            row_number=0,
                            error_code=ImportErrorCode.MISSING_SMILES_COLUMN,
                            message=f"Could not auto-detect SMILES column. "
                            f"Available columns: {file_result.headers}",
                        )
                    ],
                    error_count=1,
                    file_type=file_result.file_type,
                    detected_columns=file_result.headers,
                )
        elif isinstance(mapping, dict):
            col_mapping, mapping_errors = create_mapping(mapping, file_result.headers)
        elif isinstance(mapping, ColumnMapping):
            mapping_errors = mapping.validate(file_result.headers)
            if not mapping_errors:
                col_mapping = mapping

        if mapping_errors:
            return BatchImportResult(
                errors=[
                    RowError(
                        row_number=0,
                        error_code=ImportErrorCode.INVALID_COLUMN_MAPPING,
                        message="; ".join(mapping_errors),
                    )
                ],
                error_count=1,
                file_type=file_result.file_type,
                detected_columns=file_result.headers,
            )

        # Process rows
        return self._process_rows(
            rows=file_result.rows,
            mapping=col_mapping,
            columns=file_result.headers,
            file_type=file_result.file_type,
            file_path=str(source) if isinstance(source, (str, Path)) else None,
        )

    def _process_rows(
        self,
        rows: list[dict[str, Any]],
        mapping: ColumnMapping,
        columns: list[str],
        file_type: FileType,
        file_path: str | None,
    ) -> BatchImportResult:
        """Process rows and validate SMILES."""
        molecules = []
        errors = []

        # Get actual column names (case-matched)
        smiles_col = mapping.get_actual_column_name(mapping.smiles_col, columns) or mapping.smiles_col
        name_col = mapping.get_actual_column_name(mapping.name_col, columns) if mapping.name_col else None
        id_col = mapping.get_actual_column_name(mapping.id_col, columns) if mapping.id_col else None

        for idx, row in enumerate(rows):
            row_number = idx + 2  # 1-based, accounting for header

            # Get SMILES value
            smiles_value = row.get(smiles_col, "")
            if self.strip_whitespace:
                smiles_value = smiles_value.strip() if smiles_value else ""

            # Skip empty SMILES if configured
            if not smiles_value:
                if self.skip_empty_rows:
                    continue
                errors.append(
                    RowError(
                        row_number=row_number,
                        error_code=ImportErrorCode.EMPTY_SMILES,
                        message="Empty SMILES value",
                        column_name=smiles_col,
                    )
                )
                continue

            # Validate SMILES
            canonical_smiles = None
            inchikey = None
            is_valid = True
            warnings = []

            if self.validate_smiles or self.canonicalize:
                try:
                    from packages.chemistry.smiles import (
                        canonicalize_smiles as canon_func,
                        validate_smiles as validate_func,
                    )

                    if self.canonicalize:
                        result = canon_func(smiles_value)
                        canonical_smiles = result.canonical_smiles
                        inchikey = result.inchikey
                        if result.warnings:
                            warnings.extend(result.warnings)
                    elif self.validate_smiles:
                        is_valid = validate_func(smiles_value)
                        if not is_valid:
                            errors.append(
                                RowError(
                                    row_number=row_number,
                                    error_code=ImportErrorCode.INVALID_SMILES,
                                    message=f"Invalid SMILES: '{smiles_value}'",
                                    smiles_value=smiles_value,
                                    column_name=smiles_col,
                                )
                            )
                            continue

                except Exception as e:
                    errors.append(
                        RowError(
                            row_number=row_number,
                            error_code=ImportErrorCode.VALIDATION_FAILED,
                            message=f"Validation error: {e}",
                            smiles_value=smiles_value,
                            column_name=smiles_col,
                        )
                    )
                    continue

            # Extract optional fields
            name_value = None
            if name_col:
                name_value = row.get(name_col, "")
                if self.strip_whitespace:
                    name_value = name_value.strip() if name_value else ""
                name_value = name_value or None

            id_value = None
            if id_col:
                id_value = row.get(id_col, "")
                if self.strip_whitespace:
                    id_value = id_value.strip() if id_value else ""
                id_value = id_value or None

            # Extract extra columns
            extra_data = {}
            for key, col in mapping.extra_cols.items():
                actual_col = mapping.get_actual_column_name(col, columns) or col
                value = row.get(actual_col, "")
                if self.strip_whitespace:
                    value = value.strip() if value else ""
                if value:
                    extra_data[key] = value

            # Create molecule record
            molecules.append(
                ImportedMolecule(
                    row_number=row_number,
                    smiles=smiles_value,
                    name=name_value,
                    external_id=id_value,
                    extra_data=extra_data,
                    canonical_smiles=canonical_smiles,
                    inchikey=inchikey,
                    is_valid=is_valid,
                    warnings=warnings,
                )
            )

        return BatchImportResult(
            molecules=molecules,
            errors=errors,
            total_rows=len(rows),
            success_count=len(molecules),
            error_count=len(errors),
            file_path=file_path,
            file_type=file_type,
            detected_columns=columns,
            used_mapping=mapping,
        )


# =============================================================================
# Convenience Functions
# =============================================================================


def import_molecules_from_file(
    source: str | Path | TextIO | BinaryIO | bytes,
    mapping: dict[str, str] | ColumnMapping | None = None,
    validate: bool = True,
    canonicalize: bool = False,
    file_type: FileType | None = None,
    sheet_name: str | int = 0,
) -> BatchImportResult:
    """
    Import molecules from a CSV or Excel file.

    Args:
        source: File path, file-like object, or bytes.
        mapping: Column mapping dict (e.g., {"smiles_col": "SMILES"}).
                 Auto-detects if None.
        validate: Whether to validate SMILES strings.
        canonicalize: Whether to canonicalize SMILES (includes InChIKey).
        file_type: Explicit file type (auto-detect if None).
        sheet_name: Excel sheet name or index.

    Returns:
        BatchImportResult with molecules and errors.

    Example:
        >>> # With explicit mapping
        >>> result = import_molecules_from_file(
        ...     "compounds.csv",
        ...     mapping={"smiles_col": "SMILES", "name_col": "Name", "id_col": "CAS"}
        ... )

        >>> # With auto-detection
        >>> result = import_molecules_from_file("compounds.xlsx")

        >>> # Process results
        >>> for mol in result.molecules:
        ...     print(f"Row {mol.row_number}: {mol.smiles}")
        >>> for err in result.errors:
        ...     print(f"Error at row {err.row_number}: {err.message}")
    """
    importer = BatchImporter(
        validate_smiles=validate,
        canonicalize=canonicalize,
    )
    return importer.import_from_file(
        source=source,
        mapping=mapping,
        file_type=file_type,
        sheet_name=sheet_name,
    )


def import_molecules_from_csv(
    source: str | Path | TextIO,
    mapping: dict[str, str] | None = None,
    delimiter: str = ",",
    validate: bool = True,
) -> BatchImportResult:
    """
    Import molecules from a CSV file.

    Args:
        source: File path or file-like object.
        mapping: Column mapping dict.
        delimiter: CSV delimiter.
        validate: Whether to validate SMILES.

    Returns:
        BatchImportResult.
    """
    file_type = FileType.CSV if delimiter == "," else FileType.TSV
    return import_molecules_from_file(
        source=source,
        mapping=mapping,
        validate=validate,
        file_type=file_type,
    )


def import_molecules_from_excel(
    source: str | Path | BinaryIO | bytes,
    mapping: dict[str, str] | None = None,
    sheet_name: str | int = 0,
    validate: bool = True,
) -> BatchImportResult:
    """
    Import molecules from an Excel file.

    Args:
        source: File path, file-like object, or bytes.
        mapping: Column mapping dict.
        sheet_name: Sheet name or index.
        validate: Whether to validate SMILES.

    Returns:
        BatchImportResult.
    """
    return import_molecules_from_file(
        source=source,
        mapping=mapping,
        validate=validate,
        file_type=FileType.EXCEL,
        sheet_name=sheet_name,
    )


def validate_mapping(
    mapping: dict[str, str],
    columns: list[str],
) -> tuple[bool, list[str]]:
    """
    Validate a column mapping against available columns.

    Args:
        mapping: Mapping dict to validate.
        columns: Available column names.

    Returns:
        Tuple of (is_valid, error_messages).

    Example:
        >>> is_valid, errors = validate_mapping(
        ...     {"smiles_col": "SMILES", "name_col": "Name"},
        ...     ["SMILES", "ID", "MW"]
        ... )
        >>> is_valid
        False
        >>> errors
        ["Name column 'Name' not found"]
    """
    col_mapping, errors = create_mapping(mapping, columns)
    return (col_mapping is not None, errors)
