"""
Batch molecular data parsing from CSV and Excel files.

Supports:
- CSV files with SMILES column
- Excel files (.xlsx, .xls) with SMILES column
- Column auto-detection or explicit specification
- Metadata column extraction
"""

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO, TextIO

from packages.chemistry.exceptions import ChemistryErrorCode, ParsingError, RowError
from packages.chemistry.schemas import InputFormat, MoleculeInput

# Common SMILES column names (case-insensitive matching)
SMILES_COLUMN_NAMES = [
    "smiles",
    "canonical_smiles",
    "structure",
    "molecule",
    "mol",
    "compound",
    "smi",
    "isomeric_smiles",
]

# Common name column names
NAME_COLUMN_NAMES = [
    "name",
    "compound_name",
    "molecule_name",
    "mol_name",
    "title",
    "id",
    "compound_id",
    "mol_id",
]


@dataclass
class BatchParseResult:
    """Result of batch file parsing."""

    molecules: list[MoleculeInput]
    errors: list[RowError]
    total_rows: int
    column_mapping: dict[str, str] = field(default_factory=dict)


class BatchParser:
    """Parser for batch molecular data from CSV/Excel files."""

    def __init__(
        self,
        smiles_column: str | None = None,
        name_column: str | None = None,
        metadata_columns: list[str] | None = None,
        skip_header: bool = True,
    ):
        """
        Initialize batch parser.

        Args:
            smiles_column: Name of SMILES column (auto-detect if None).
            name_column: Name of molecule name column (auto-detect if None).
            metadata_columns: Additional columns to include as metadata.
            skip_header: Whether first row is header.
        """
        self.smiles_column = smiles_column
        self.name_column = name_column
        self.metadata_columns = metadata_columns or []
        self.skip_header = skip_header

    def _find_column(
        self, headers: list[str], candidates: list[str]
    ) -> str | None:
        """Find matching column from candidates (case-insensitive)."""
        headers_lower = {h.lower().strip(): h for h in headers}
        for candidate in candidates:
            if candidate.lower() in headers_lower:
                return headers_lower[candidate.lower()]
        return None

    def _detect_columns(
        self, headers: list[str]
    ) -> tuple[str | None, str | None]:
        """Auto-detect SMILES and name columns from headers."""
        smiles_col = self.smiles_column or self._find_column(
            headers, SMILES_COLUMN_NAMES
        )
        name_col = self.name_column or self._find_column(headers, NAME_COLUMN_NAMES)
        return smiles_col, name_col

    def parse_csv(
        self,
        content: str | TextIO,
        delimiter: str = ",",
    ) -> BatchParseResult:
        """
        Parse molecules from CSV content.

        Args:
            content: CSV string or file-like object.
            delimiter: Column delimiter.

        Returns:
            BatchParseResult with parsed molecules and errors.

        Raises:
            ParsingError: If CSV format is invalid.
        """
        if isinstance(content, str):
            content = io.StringIO(content)

        try:
            reader = csv.DictReader(content, delimiter=delimiter)
            headers = reader.fieldnames or []
        except Exception as e:
            raise ParsingError(
                message=f"Failed to parse CSV: {e}",
                code=ChemistryErrorCode.UNSUPPORTED_FORMAT,
            )

        if not headers:
            raise ParsingError(
                message="CSV has no headers",
                code=ChemistryErrorCode.EMPTY_INPUT,
            )

        smiles_col, name_col = self._detect_columns(headers)

        if not smiles_col:
            raise ParsingError(
                message=f"SMILES column not found. Headers: {headers}",
                code=ChemistryErrorCode.UNSUPPORTED_FORMAT,
                details={"headers": headers},
            )

        molecules = []
        errors = []
        row_index = 0

        for row in reader:
            row_index += 1
            smiles = row.get(smiles_col, "").strip()

            if not smiles:
                errors.append(
                    RowError(
                        row_index=row_index,
                        input_value="",
                        error_code=ChemistryErrorCode.EMPTY_INPUT,
                        error_message=f"Empty SMILES in row {row_index}",
                    )
                )
                continue

            # Extract name
            name = row.get(name_col, "").strip() if name_col else None

            # Extract metadata from specified columns
            metadata = {}
            for col in self.metadata_columns:
                if col in row and col not in (smiles_col, name_col):
                    metadata[col] = row[col]

            molecules.append(
                MoleculeInput(
                    value=smiles,
                    format=InputFormat.SMILES,
                    name=name or None,
                    metadata=metadata if metadata else None,
                )
            )

        return BatchParseResult(
            molecules=molecules,
            errors=errors,
            total_rows=row_index,
            column_mapping={
                "smiles": smiles_col,
                "name": name_col,
            },
        )

    def parse_excel(
        self,
        content: bytes | BinaryIO | Path,
        sheet_name: str | int = 0,
    ) -> BatchParseResult:
        """
        Parse molecules from Excel file.

        Args:
            content: Excel file content, file-like object, or path.
            sheet_name: Sheet name or index (0-based).

        Returns:
            BatchParseResult with parsed molecules and errors.

        Raises:
            ParsingError: If Excel file is invalid.
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel parsing. "
                "Install with: pip install openpyxl"
            )

        try:
            if isinstance(content, Path):
                wb = openpyxl.load_workbook(content, read_only=True, data_only=True)
            elif isinstance(content, bytes):
                wb = openpyxl.load_workbook(
                    io.BytesIO(content), read_only=True, data_only=True
                )
            else:
                wb = openpyxl.load_workbook(content, read_only=True, data_only=True)
        except Exception as e:
            raise ParsingError(
                message=f"Failed to open Excel file: {e}",
                code=ChemistryErrorCode.UNSUPPORTED_FORMAT,
            )

        # Get sheet
        if isinstance(sheet_name, int):
            if sheet_name >= len(wb.sheetnames):
                raise ParsingError(
                    message=f"Sheet index {sheet_name} out of range",
                    code=ChemistryErrorCode.UNSUPPORTED_FORMAT,
                )
            ws = wb[wb.sheetnames[sheet_name]]
        else:
            if sheet_name not in wb.sheetnames:
                raise ParsingError(
                    message=f"Sheet '{sheet_name}' not found",
                    code=ChemistryErrorCode.UNSUPPORTED_FORMAT,
                )
            ws = wb[sheet_name]

        # Get headers from first row
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ParsingError(
                message="Excel sheet is empty",
                code=ChemistryErrorCode.EMPTY_INPUT,
            )

        headers = [str(h).strip() if h else "" for h in header_row]
        smiles_col, name_col = self._detect_columns(headers)

        if not smiles_col:
            raise ParsingError(
                message=f"SMILES column not found. Headers: {headers}",
                code=ChemistryErrorCode.UNSUPPORTED_FORMAT,
                details={"headers": headers},
            )

        smiles_idx = headers.index(smiles_col)
        name_idx = headers.index(name_col) if name_col else None

        # Build metadata column indices
        metadata_indices = {}
        for col in self.metadata_columns:
            if col in headers and col not in (smiles_col, name_col):
                metadata_indices[col] = headers.index(col)

        molecules = []
        errors = []
        row_index = 1  # 1-based (header is row 1)

        for row in rows_iter:
            row_index += 1

            # Get SMILES
            smiles = str(row[smiles_idx]).strip() if row[smiles_idx] else ""

            if not smiles:
                errors.append(
                    RowError(
                        row_index=row_index,
                        input_value="",
                        error_code=ChemistryErrorCode.EMPTY_INPUT,
                        error_message=f"Empty SMILES in row {row_index}",
                    )
                )
                continue

            # Get name
            name = None
            if name_idx is not None and row[name_idx]:
                name = str(row[name_idx]).strip()

            # Get metadata
            metadata = {}
            for col, idx in metadata_indices.items():
                if idx < len(row) and row[idx]:
                    metadata[col] = str(row[idx])

            molecules.append(
                MoleculeInput(
                    value=smiles,
                    format=InputFormat.SMILES,
                    name=name or None,
                    metadata=metadata if metadata else None,
                )
            )

        wb.close()

        return BatchParseResult(
            molecules=molecules,
            errors=errors,
            total_rows=row_index - 1,  # Exclude header
            column_mapping={
                "smiles": smiles_col,
                "name": name_col,
            },
        )


def parse_csv(
    content: str | TextIO,
    smiles_column: str | None = None,
    delimiter: str = ",",
) -> BatchParseResult:
    """
    Convenience function to parse CSV content.

    Args:
        content: CSV string or file-like object.
        smiles_column: Name of SMILES column (auto-detect if None).
        delimiter: Column delimiter.

    Returns:
        BatchParseResult.
    """
    parser = BatchParser(smiles_column=smiles_column)
    return parser.parse_csv(content, delimiter=delimiter)


def parse_excel(
    content: bytes | BinaryIO | Path,
    smiles_column: str | None = None,
    sheet_name: str | int = 0,
) -> BatchParseResult:
    """
    Convenience function to parse Excel content.

    Args:
        content: Excel file content, file-like object, or path.
        smiles_column: Name of SMILES column (auto-detect if None).
        sheet_name: Sheet name or index.

    Returns:
        BatchParseResult.
    """
    parser = BatchParser(smiles_column=smiles_column)
    return parser.parse_excel(content, sheet_name=sheet_name)
