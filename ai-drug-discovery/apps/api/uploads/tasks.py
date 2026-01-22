"""
Background tasks for upload processing.

Handles:
- File parsing (SDF, CSV, Excel, SMILES list)
- Column mapping inference for CSV/Excel
- Molecule validation
- Duplicate detection
- Database insertion
"""

import asyncio
import csv
import hashlib
import logging
import time
import uuid
from decimal import Decimal
from io import BytesIO, StringIO
from typing import AsyncIterator, NamedTuple

from sqlalchemy.ext.asyncio import AsyncSession

from apps.api.uploads.error_codes import UploadErrorCode
from apps.api.uploads.file_detection import (
    detect_csv_columns,
    detect_excel_columns,
    infer_column_mapping,
)
from apps.api.uploads.service import UploadService
from db.models.discovery import Molecule
from db.models.upload import DuplicateAction, FileType, Upload, UploadRowError

logger = logging.getLogger(__name__)

# Import chemistry utilities
try:
    from packages.chemistry import (
        calculate_descriptors_rdkit,
        calculate_morgan_fingerprint,
        canonicalize_smiles,
        validate_smiles,
    )
    from packages.chemistry.smiles import smiles_to_mol
    from rdkit import Chem
    from rdkit.Chem.inchi import MolToInchi, MolToInchiKey

    RDKIT_AVAILABLE = True
except ImportError:
    RDKIT_AVAILABLE = False


class ParsedRow(NamedTuple):
    """A parsed row from the upload file."""

    row_number: int
    smiles: str
    name: str | None
    external_id: str | None
    raw_data: dict


class ValidationResult(NamedTuple):
    """Result of validating a single molecule."""

    row_number: int
    is_valid: bool
    canonical_smiles: str | None
    inchi: str | None
    inchi_key: str | None
    smiles_hash: str | None
    mol: object | None  # RDKit Mol
    error_code: UploadErrorCode | None
    error_detail: str | None
    raw_data: dict


class UploadProcessor:
    """
    Background processor for upload validation and molecule insertion.

    Handles the full lifecycle of processing an upload file.
    """

    # Batch sizes
    PARSE_BATCH_SIZE = 100
    VALIDATE_BATCH_SIZE = 50
    INSERT_BATCH_SIZE = 100
    PROGRESS_UPDATE_INTERVAL = 50

    def __init__(
        self,
        db: AsyncSession,
        service: UploadService,
    ):
        """
        Initialize processor.

        Args:
            db: Database session
            service: Upload service instance
        """
        self.db = db
        self.service = service

    # =========================================================================
    # Main Processing Entry Points
    # =========================================================================

    async def process_validation(self, upload: Upload) -> None:
        """
        Run validation on an upload.

        This is called as a background task after upload creation.
        Transitions upload from INITIATED -> VALIDATING -> AWAITING_CONFIRM/VALIDATION_FAILED.

        For CSV/Excel files without column mapping:
        - Detects available columns
        - Attempts to infer mapping
        - If SMILES column cannot be inferred, moves to AWAITING_CONFIRM with needs_column_mapping=True

        Args:
            upload: Upload to validate
        """
        try:
            # Start validation
            await self.service.start_validation(upload)

            # Get file content
            file_content = await self.service.get_upload_file_content(upload)

            # For CSV/Excel, check column mapping
            if upload.file_type in (FileType.CSV, FileType.EXCEL):
                needs_mapping = await self._check_column_mapping(upload, file_content)
                if needs_mapping:
                    # Move to AWAITING_CONFIRM with needs_mapping flag
                    await self.service.complete_validation_needs_mapping(upload)
                    return
                # Reset file position after column check
                file_content.seek(0)

            # Parse and validate
            total_rows = 0
            valid_rows = 0
            invalid_rows = 0
            duplicate_exact = 0
            duplicate_similar = 0
            seen_inchi_keys: set[str] = set()  # Track duplicates within batch

            # Parse file based on type
            async for batch in self._parse_file(upload, file_content):
                # Validate batch
                for result in await self._validate_batch(upload, batch, seen_inchi_keys):
                    total_rows += 1

                    if result.is_valid:
                        # Check for duplicates
                        dup_type = await self._check_duplicates(
                            upload, result, seen_inchi_keys
                        )
                        if dup_type == "exact":
                            duplicate_exact += 1
                            if upload.duplicate_action == DuplicateAction.ERROR:
                                invalid_rows += 1
                            else:
                                valid_rows += 1
                        elif dup_type == "similar":
                            duplicate_similar += 1
                            if upload.duplicate_action == DuplicateAction.ERROR:
                                invalid_rows += 1
                            else:
                                valid_rows += 1
                        elif dup_type == "batch":
                            # Duplicate within this upload
                            invalid_rows += 1
                        else:
                            valid_rows += 1
                            if result.inchi_key:
                                seen_inchi_keys.add(result.inchi_key)
                    else:
                        invalid_rows += 1

                    # Update progress periodically
                    if total_rows % self.PROGRESS_UPDATE_INTERVAL == 0:
                        await self.service.update_progress(
                            upload,
                            processed_rows=total_rows,
                            valid_rows=valid_rows,
                            invalid_rows=invalid_rows,
                            duplicate_exact=duplicate_exact,
                            duplicate_similar=duplicate_similar,
                            phase="validating",
                        )

            # Complete validation
            await self.service.complete_validation(
                upload,
                total_rows=total_rows,
                valid_rows=valid_rows,
                invalid_rows=invalid_rows,
                duplicate_exact=duplicate_exact,
                duplicate_similar=duplicate_similar,
            )

        except Exception as e:
            await self.service.fail_upload(upload, str(e))
            raise

    async def process_insertion(self, upload: Upload) -> None:
        """
        Insert validated molecules into the database.

        Called after user confirms the upload.
        Transitions upload from PROCESSING -> COMPLETED/FAILED.

        Args:
            upload: Confirmed upload to process
        """
        start_time = time.time()

        try:
            # Get file content
            file_content = await self.service.get_upload_file_content(upload)

            # Track results
            molecules_created = 0
            molecules_updated = 0
            molecules_skipped = 0
            errors_count = 0
            exact_duplicates = 0
            similar_duplicates = 0
            seen_inchi_keys: set[str] = set()
            processed_rows = 0

            # Process file
            async for batch in self._parse_file(upload, file_content):
                for result in await self._validate_batch(upload, batch, seen_inchi_keys):
                    processed_rows += 1

                    if not result.is_valid:
                        errors_count += 1
                        continue

                    # Check duplicates
                    dup_type = await self._check_duplicates(
                        upload, result, seen_inchi_keys, record_errors=False
                    )

                    if dup_type == "batch":
                        # Skip duplicates within batch
                        molecules_skipped += 1
                        continue

                    if dup_type == "exact":
                        exact_duplicates += 1
                        if upload.duplicate_action == DuplicateAction.SKIP:
                            molecules_skipped += 1
                            continue
                        elif upload.duplicate_action == DuplicateAction.UPDATE:
                            await self._update_molecule(upload, result)
                            molecules_updated += 1
                            continue
                        else:  # ERROR - should have been caught in validation
                            errors_count += 1
                            continue

                    if dup_type == "similar":
                        similar_duplicates += 1
                        if upload.duplicate_action in (DuplicateAction.SKIP, DuplicateAction.UPDATE):
                            molecules_skipped += 1
                            continue
                        else:
                            errors_count += 1
                            continue

                    # Insert new molecule
                    try:
                        await self._insert_molecule(upload, result)
                        molecules_created += 1
                        if result.inchi_key:
                            seen_inchi_keys.add(result.inchi_key)
                    except Exception as e:
                        errors_count += 1
                        await self.service.add_row_error(
                            upload,
                            result.row_number,
                            UploadErrorCode.DB_INSERT_FAILED,
                            str(e),
                            raw_data=result.raw_data,
                        )

                    # Update progress
                    if processed_rows % self.PROGRESS_UPDATE_INTERVAL == 0:
                        await self.service.update_progress(
                            upload,
                            processed_rows=processed_rows,
                            phase="inserting",
                        )

            # Complete
            duration = time.time() - start_time
            await self.service.complete_processing(
                upload,
                molecules_created=molecules_created,
                molecules_updated=molecules_updated,
                molecules_skipped=molecules_skipped,
                errors_count=errors_count,
                exact_duplicates=exact_duplicates,
                similar_duplicates=similar_duplicates,
                duration_seconds=duration,
            )

        except Exception as e:
            await self.service.fail_upload(upload, str(e))
            raise

    # =========================================================================
    # Column Mapping Check (CSV/Excel)
    # =========================================================================

    async def _check_column_mapping(
        self,
        upload: Upload,
        file_content: BytesIO,
    ) -> bool:
        """
        Check if column mapping is needed for CSV/Excel files.

        If mapping is not provided, attempts to infer it.
        Updates upload record with available columns and inferred mapping.

        Args:
            upload: Upload record
            file_content: File content

        Returns:
            True if user needs to provide mapping, False if we can proceed
        """
        # If mapping already provided, we're good
        if upload.column_mapping and upload.column_mapping.get("smiles"):
            return False

        # Detect columns from file
        content = file_content.read()
        file_content.seek(0)

        if upload.file_type == FileType.CSV:
            columns = detect_csv_columns(content)
        elif upload.file_type == FileType.EXCEL:
            columns = detect_excel_columns(content)
        else:
            return False

        if not columns:
            # Can't detect columns - needs manual mapping
            upload.needs_column_mapping = True
            upload.available_columns = []
            await self.db.commit()
            return True

        # Store available columns
        upload.available_columns = columns

        # Try to infer mapping
        inferred = infer_column_mapping(columns)
        upload.inferred_mapping = inferred

        if inferred.get("smiles"):
            # We found a SMILES column - use inferred mapping
            upload.column_mapping = inferred
            upload.needs_column_mapping = False
            await self.db.commit()
            logger.info(f"Auto-inferred column mapping for upload {upload.id}: {inferred}")
            return False
        else:
            # Could not find SMILES column - user needs to provide mapping
            upload.needs_column_mapping = True
            await self.db.commit()
            logger.info(f"Upload {upload.id} needs column mapping. Columns: {columns}")
            return True

    # =========================================================================
    # File Parsing
    # =========================================================================

    async def _parse_file(
        self,
        upload: Upload,
        file_content: BytesIO,
    ) -> AsyncIterator[list[ParsedRow]]:
        """
        Parse upload file and yield batches of rows.

        Args:
            upload: Upload record
            file_content: File content

        Yields:
            Batches of ParsedRow objects
        """
        if upload.file_type == FileType.CSV:
            async for batch in self._parse_csv(upload, file_content):
                yield batch
        elif upload.file_type == FileType.EXCEL:
            async for batch in self._parse_excel(upload, file_content):
                yield batch
        elif upload.file_type == FileType.SDF:
            async for batch in self._parse_sdf(upload, file_content):
                yield batch
        elif upload.file_type == FileType.SMILES_LIST:
            async for batch in self._parse_smiles_list(upload, file_content):
                yield batch
        else:
            raise ValueError(f"Unsupported file type: {upload.file_type}")

    async def _parse_csv(
        self,
        upload: Upload,
        file_content: BytesIO,
    ) -> AsyncIterator[list[ParsedRow]]:
        """Parse CSV file."""
        mapping = upload.column_mapping or {}
        smiles_col = mapping.get("smiles", "smiles")
        name_col = mapping.get("name")
        external_id_col = mapping.get("external_id")

        # Decode content
        text_content = file_content.read().decode("utf-8")
        reader = csv.DictReader(StringIO(text_content))

        batch: list[ParsedRow] = []
        row_number = 1  # 1-based, row 1 is header

        for row in reader:
            row_number += 1
            smiles = row.get(smiles_col, "").strip()

            parsed = ParsedRow(
                row_number=row_number,
                smiles=smiles,
                name=row.get(name_col, "").strip() if name_col else None,
                external_id=row.get(external_id_col, "").strip() if external_id_col else None,
                raw_data=dict(row),
            )
            batch.append(parsed)

            if len(batch) >= self.PARSE_BATCH_SIZE:
                yield batch
                batch = []
                await asyncio.sleep(0)  # Yield control

        if batch:
            yield batch

    async def _parse_excel(
        self,
        upload: Upload,
        file_content: BytesIO,
    ) -> AsyncIterator[list[ParsedRow]]:
        """Parse Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")

        mapping = upload.column_mapping or {}
        smiles_col = mapping.get("smiles", "smiles")
        name_col = mapping.get("name")
        external_id_col = mapping.get("external_id")

        # Load workbook
        content = file_content.read()
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
        ws = wb.active

        if not ws:
            return

        # Get headers from first row
        headers: list[str] = []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            headers = [str(cell) if cell else "" for cell in header_row]

        # Find column indices
        try:
            smiles_idx = headers.index(smiles_col)
        except ValueError:
            # Try case-insensitive match
            smiles_idx = None
            for i, h in enumerate(headers):
                if h.lower() == smiles_col.lower():
                    smiles_idx = i
                    break
            if smiles_idx is None:
                raise ValueError(f"SMILES column '{smiles_col}' not found in Excel headers: {headers}")

        name_idx = None
        if name_col:
            for i, h in enumerate(headers):
                if h.lower() == name_col.lower():
                    name_idx = i
                    break

        external_id_idx = None
        if external_id_col:
            for i, h in enumerate(headers):
                if h.lower() == external_id_col.lower():
                    external_id_idx = i
                    break

        batch: list[ParsedRow] = []
        row_number = 1  # 1-based, row 1 is header

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_number += 1
            row_values = list(row)

            # Get SMILES
            smiles = ""
            if smiles_idx < len(row_values) and row_values[smiles_idx]:
                smiles = str(row_values[smiles_idx]).strip()

            # Get name
            name = None
            if name_idx is not None and name_idx < len(row_values) and row_values[name_idx]:
                name = str(row_values[name_idx]).strip()

            # Get external ID
            external_id = None
            if external_id_idx is not None and external_id_idx < len(row_values) and row_values[external_id_idx]:
                external_id = str(row_values[external_id_idx]).strip()

            # Build raw_data dict
            raw_data = {headers[i]: str(v) if v else "" for i, v in enumerate(row_values) if i < len(headers)}

            parsed = ParsedRow(
                row_number=row_number,
                smiles=smiles,
                name=name,
                external_id=external_id,
                raw_data=raw_data,
            )
            batch.append(parsed)

            if len(batch) >= self.PARSE_BATCH_SIZE:
                yield batch
                batch = []
                await asyncio.sleep(0)  # Yield control

        if batch:
            yield batch

        wb.close()

    async def _parse_sdf(
        self,
        upload: Upload,
        file_content: BytesIO,
    ) -> AsyncIterator[list[ParsedRow]]:
        """Parse SDF file."""
        if not RDKIT_AVAILABLE:
            raise ImportError("RDKit is required for SDF parsing")

        content = file_content.read()
        suppl = Chem.SDMolSupplier()
        suppl.SetData(content.decode("utf-8"))

        batch: list[ParsedRow] = []
        row_number = 0

        for mol in suppl:
            row_number += 1

            if mol is None:
                # Invalid molecule in SDF
                batch.append(ParsedRow(
                    row_number=row_number,
                    smiles="",  # Will fail validation
                    name=None,
                    external_id=None,
                    raw_data={"error": "Failed to parse molecule from SDF"},
                ))
            else:
                smiles = Chem.MolToSmiles(mol)
                name = mol.GetProp("_Name") if mol.HasProp("_Name") else None

                # Get all properties as raw_data
                raw_data = {prop: mol.GetProp(prop) for prop in mol.GetPropsAsDict()}

                batch.append(ParsedRow(
                    row_number=row_number,
                    smiles=smiles,
                    name=name,
                    external_id=raw_data.get("CAS", raw_data.get("external_id")),
                    raw_data=raw_data,
                ))

            if len(batch) >= self.PARSE_BATCH_SIZE:
                yield batch
                batch = []
                await asyncio.sleep(0)

        if batch:
            yield batch

    async def _parse_smiles_list(
        self,
        upload: Upload,
        file_content: BytesIO,
    ) -> AsyncIterator[list[ParsedRow]]:
        """Parse SMILES list (one per line, optional tab-separated name)."""
        text_content = file_content.read().decode("utf-8")
        lines = text_content.strip().split("\n")

        batch: list[ParsedRow] = []

        for row_number, line in enumerate(lines, start=1):
            line = line.strip()
            if not line or line.startswith("#"):
                continue

            # Split by tab or whitespace
            parts = line.split("\t") if "\t" in line else line.split(None, 1)
            smiles = parts[0].strip()
            name = parts[1].strip() if len(parts) > 1 else None

            batch.append(ParsedRow(
                row_number=row_number,
                smiles=smiles,
                name=name,
                external_id=None,
                raw_data={"smiles": smiles, "name": name},
            ))

            if len(batch) >= self.PARSE_BATCH_SIZE:
                yield batch
                batch = []
                await asyncio.sleep(0)

        if batch:
            yield batch

    # =========================================================================
    # Validation
    # =========================================================================

    async def _validate_batch(
        self,
        upload: Upload,
        batch: list[ParsedRow],
        seen_inchi_keys: set[str],
    ) -> list[ValidationResult]:
        """
        Validate a batch of parsed rows.

        Args:
            upload: Upload record
            batch: Parsed rows to validate
            seen_inchi_keys: InChIKeys already seen in this upload

        Returns:
            List of ValidationResult objects
        """
        results: list[ValidationResult] = []
        errors_to_add: list[UploadRowError] = []

        for row in batch:
            result = await self._validate_molecule(row)
            results.append(result)

            if not result.is_valid and result.error_code:
                error = UploadRowError(
                    upload_id=upload.id,
                    row_number=row.row_number,
                    error_code=result.error_code.value,
                    error_message=result.error_detail or "",
                    raw_data=result.raw_data,
                )
                errors_to_add.append(error)

        if errors_to_add:
            await self.service.add_row_errors_batch(errors_to_add)

        return results

    async def _validate_molecule(self, row: ParsedRow) -> ValidationResult:
        """
        Validate a single molecule.

        Args:
            row: Parsed row

        Returns:
            ValidationResult
        """
        if not RDKIT_AVAILABLE:
            # Minimal validation without RDKit
            return await self._validate_molecule_minimal(row)

        smiles = row.smiles

        # Check empty
        if not smiles:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=None,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=None,
                error_code=UploadErrorCode.MISSING_REQUIRED_FIELD,
                error_detail="SMILES is empty",
                raw_data=row.raw_data,
            )

        # Check length
        if len(smiles) > 2000:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=None,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=None,
                error_code=UploadErrorCode.SMILES_TOO_LONG,
                error_detail=f"SMILES length {len(smiles)} exceeds 2000",
                raw_data=row.raw_data,
            )

        # Parse with RDKit
        try:
            mol = smiles_to_mol(smiles)
            if mol is None:
                return ValidationResult(
                    row_number=row.row_number,
                    is_valid=False,
                    canonical_smiles=None,
                    inchi=None,
                    inchi_key=None,
                    smiles_hash=None,
                    mol=None,
                    error_code=UploadErrorCode.INVALID_SMILES,
                    error_detail=f"Cannot parse SMILES: {smiles[:50]}",
                    raw_data=row.raw_data,
                )
        except Exception as e:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=None,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=None,
                error_code=UploadErrorCode.INVALID_SMILES,
                error_detail=str(e),
                raw_data=row.raw_data,
            )

        # Check molecule size
        num_atoms = mol.GetNumHeavyAtoms()
        if num_atoms > 1000:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=None,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=mol,
                error_code=UploadErrorCode.MOLECULE_TOO_LARGE,
                error_detail=f"Molecule has {num_atoms} heavy atoms (max 1000)",
                raw_data=row.raw_data,
            )

        if num_atoms == 0:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=None,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=mol,
                error_code=UploadErrorCode.NO_ATOMS,
                error_detail="Molecule has no atoms",
                raw_data=row.raw_data,
            )

        # Generate canonical SMILES
        try:
            canonical = Chem.MolToSmiles(mol, canonical=True)
        except Exception:
            canonical = smiles

        # Generate InChI and InChIKey
        try:
            inchi = MolToInchi(mol)
            inchi_key = MolToInchiKey(mol)
        except Exception as e:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=canonical,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=mol,
                error_code=UploadErrorCode.INCHI_GENERATION_FAILED,
                error_detail=str(e),
                raw_data=row.raw_data,
            )

        # Generate SMILES hash
        smiles_hash = hashlib.sha256(canonical.encode()).hexdigest()

        return ValidationResult(
            row_number=row.row_number,
            is_valid=True,
            canonical_smiles=canonical,
            inchi=inchi,
            inchi_key=inchi_key,
            smiles_hash=smiles_hash,
            mol=mol,
            error_code=None,
            error_detail=None,
            raw_data=row.raw_data,
        )

    async def _validate_molecule_minimal(self, row: ParsedRow) -> ValidationResult:
        """Minimal validation without RDKit."""
        smiles = row.smiles

        if not smiles:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=None,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=None,
                error_code=UploadErrorCode.MISSING_REQUIRED_FIELD,
                error_detail="SMILES is empty",
                raw_data=row.raw_data,
            )

        if len(smiles) > 2000:
            return ValidationResult(
                row_number=row.row_number,
                is_valid=False,
                canonical_smiles=None,
                inchi=None,
                inchi_key=None,
                smiles_hash=None,
                mol=None,
                error_code=UploadErrorCode.SMILES_TOO_LONG,
                error_detail=f"SMILES length {len(smiles)} exceeds 2000",
                raw_data=row.raw_data,
            )

        # Generate SMILES hash (use original as canonical without RDKit)
        smiles_hash = hashlib.sha256(smiles.encode()).hexdigest()

        return ValidationResult(
            row_number=row.row_number,
            is_valid=True,
            canonical_smiles=smiles,
            inchi=None,
            inchi_key=f"PLACEHOLDER-{smiles_hash[:27]}",  # Fake InChIKey
            smiles_hash=smiles_hash,
            mol=None,
            error_code=None,
            error_detail=None,
            raw_data=row.raw_data,
        )

    # =========================================================================
    # Duplicate Detection
    # =========================================================================

    async def _check_duplicates(
        self,
        upload: Upload,
        result: ValidationResult,
        seen_inchi_keys: set[str],
        record_errors: bool = True,
    ) -> str | None:
        """
        Check for duplicate molecules.

        Args:
            upload: Upload record
            result: Validation result
            seen_inchi_keys: InChIKeys seen in this upload
            record_errors: Whether to record errors

        Returns:
            "exact", "similar", "batch" (duplicate in batch), or None
        """
        if not result.inchi_key:
            return None

        # Check for duplicate within batch
        if result.inchi_key in seen_inchi_keys:
            if record_errors and upload.duplicate_action == DuplicateAction.ERROR:
                await self.service.add_row_error(
                    upload,
                    result.row_number,
                    UploadErrorCode.DUPLICATE_IN_BATCH,
                    f"Duplicate of another row in this upload",
                    raw_data=result.raw_data,
                    duplicate_inchi_key=result.inchi_key,
                )
            return "batch"

        # Check for exact duplicate in database
        existing = await self.service.check_exact_duplicate(
            upload.organization_id,
            result.inchi_key,
        )
        if existing:
            if record_errors and upload.duplicate_action == DuplicateAction.ERROR:
                await self.service.add_row_error(
                    upload,
                    result.row_number,
                    UploadErrorCode.EXACT_DUPLICATE,
                    f"Molecule already exists",
                    raw_data=result.raw_data,
                    duplicate_inchi_key=result.inchi_key,
                )
            return "exact"

        # Check for similar molecules (if threshold set and fingerprint available)
        if upload.similarity_threshold and result.mol and RDKIT_AVAILABLE:
            try:
                fp = calculate_morgan_fingerprint(result.mol)
                similar = await self.service.find_similar_molecules(
                    upload.organization_id,
                    fp.to_bytes(),
                    upload.similarity_threshold,
                    limit=1,
                )
                if similar:
                    mol, similarity = similar[0]
                    if record_errors and upload.duplicate_action == DuplicateAction.ERROR:
                        await self.service.add_row_error(
                            upload,
                            result.row_number,
                            UploadErrorCode.SIMILAR_DUPLICATE,
                            f"Similar molecule found (Tanimoto={similarity:.2f})",
                            raw_data=result.raw_data,
                            duplicate_inchi_key=mol.inchi_key,
                            duplicate_similarity=Decimal(str(similarity)),
                        )
                    return "similar"
            except Exception:
                pass  # Skip similarity check on error

        return None

    # =========================================================================
    # Database Operations
    # =========================================================================

    async def _insert_molecule(
        self,
        upload: Upload,
        result: ValidationResult,
    ) -> Molecule:
        """
        Insert a new molecule into the database.

        Stores:
        - Chemical identifiers (SMILES, InChI, InChIKey)
        - Computed descriptors (MW, LogP, HBD, HBA, TPSA, etc.)
        - Fingerprints (Morgan, MACCS, RDKit)
        - Provenance metadata (upload source, external ID)

        Args:
            upload: Upload record
            result: Validation result

        Returns:
            Created Molecule
        """
        # Calculate descriptors if RDKit available
        descriptors = {}
        fingerprints = {}

        if RDKIT_AVAILABLE and result.mol:
            try:
                desc = calculate_descriptors_rdkit(result.mol)
                descriptors = {
                    "molecular_weight": desc.molecular_weight,
                    "logp": desc.logp,
                    "hbd": desc.num_h_donors,
                    "hba": desc.num_h_acceptors,
                    "tpsa": desc.tpsa,
                    "rotatable_bonds": desc.num_rotatable_bonds,
                    "num_rings": desc.num_rings,
                    "num_aromatic_rings": desc.num_aromatic_rings,
                    "num_heavy_atoms": desc.num_heavy_atoms,
                    "fraction_sp3": desc.fraction_csp3,
                }
            except Exception as e:
                logger.warning(f"Failed to calculate descriptors for row {result.row_number}: {e}")

            # Calculate fingerprints
            try:
                fp = calculate_morgan_fingerprint(result.mol)
                fingerprints["fingerprint_morgan"] = fp.to_bytes()
            except Exception as e:
                logger.warning(f"Failed to calculate Morgan fingerprint for row {result.row_number}: {e}")

            try:
                from rdkit.Chem import MACCSkeys
                maccs = MACCSkeys.GenMACCSKeys(result.mol)
                fingerprints["fingerprint_maccs"] = maccs.ToBitString().encode()
            except Exception:
                pass

            try:
                from rdkit.Chem import RDKFingerprint
                rdkit_fp = RDKFingerprint(result.mol)
                fingerprints["fingerprint_rdkit"] = rdkit_fp.ToBitString().encode()
            except Exception:
                pass

        # Build metadata with provenance
        metadata = {
            "source_upload_id": str(upload.id),
            "source_upload_name": upload.name,
            "source_row_number": result.row_number,
        }

        # Add external_id if present
        external_id = (
            result.raw_data.get("external_id")
            or result.raw_data.get("External_ID")
            or result.raw_data.get("CAS")
            or result.raw_data.get("cas")
        )
        if external_id:
            metadata["external_id"] = external_id

        # Get name from various possible field names
        name = (
            result.raw_data.get("name")
            or result.raw_data.get("Name")
            or result.raw_data.get("compound_name")
            or result.raw_data.get("Compound_Name")
        )

        # Create molecule
        molecule = Molecule(
            organization_id=upload.organization_id,
            created_by=upload.created_by,
            canonical_smiles=result.canonical_smiles,
            inchi=result.inchi,
            inchi_key=result.inchi_key,
            smiles_hash=result.smiles_hash,
            name=name,
            metadata_=metadata,
            **descriptors,
            **fingerprints,
        )
        self.db.add(molecule)
        await self.db.flush()

        return molecule

    async def _update_molecule(
        self,
        upload: Upload,
        result: ValidationResult,
    ) -> Molecule | None:
        """
        Update an existing molecule (upsert logic).

        Updates:
        - Name (if provided and not already set)
        - Metadata with new upload provenance
        - Computed properties if missing
        - Fingerprints if missing

        Args:
            upload: Upload record
            result: Validation result

        Returns:
            Updated Molecule or None
        """
        existing = await self.service.check_exact_duplicate(
            upload.organization_id,
            result.inchi_key,
        )
        if not existing:
            return None

        # Update name if provided and molecule has no name
        name = (
            result.raw_data.get("name")
            or result.raw_data.get("Name")
            or result.raw_data.get("compound_name")
        )
        if name and not existing.name:
            existing.name = name

        # Update metadata with upload provenance
        metadata = existing.metadata_ or {}
        if "upload_history" not in metadata:
            metadata["upload_history"] = []
        metadata["upload_history"].append({
            "upload_id": str(upload.id),
            "upload_name": upload.name,
            "row_number": result.row_number,
            "action": "update",
        })
        existing.metadata_ = metadata

        # Update computed properties if missing and RDKit available
        if RDKIT_AVAILABLE and result.mol:
            # Update descriptors if missing
            if existing.molecular_weight is None:
                try:
                    desc = calculate_descriptors_rdkit(result.mol)
                    existing.molecular_weight = desc.molecular_weight
                    existing.logp = desc.logp
                    existing.hbd = desc.num_h_donors
                    existing.hba = desc.num_h_acceptors
                    existing.tpsa = desc.tpsa
                    existing.rotatable_bonds = desc.num_rotatable_bonds
                    existing.num_rings = desc.num_rings
                    existing.num_aromatic_rings = desc.num_aromatic_rings
                    existing.num_heavy_atoms = desc.num_heavy_atoms
                    existing.fraction_sp3 = desc.fraction_csp3
                except Exception:
                    pass

            # Update fingerprints if missing
            if existing.fingerprint_morgan is None:
                try:
                    fp = calculate_morgan_fingerprint(result.mol)
                    existing.fingerprint_morgan = fp.to_bytes()
                except Exception:
                    pass

        existing.updated_by = upload.created_by
        await self.db.flush()

        return existing


# =============================================================================
# Background Task Functions
# =============================================================================


async def run_validation_task(
    db: AsyncSession,
    service: UploadService,
    upload_id: uuid.UUID,
    organization_id: uuid.UUID,
) -> None:
    """
    Background task to run validation.

    Args:
        db: Database session
        service: Upload service
        upload_id: Upload ID to validate
        organization_id: Organization ID
    """
    upload = await service.get_upload(upload_id, organization_id)
    if not upload:
        return

    processor = UploadProcessor(db, service)
    await processor.process_validation(upload)


async def run_insertion_task(
    db: AsyncSession,
    service: UploadService,
    upload_id: uuid.UUID,
    organization_id: uuid.UUID,
) -> None:
    """
    Background task to run insertion.

    Args:
        db: Database session
        service: Upload service
        upload_id: Upload ID to process
        organization_id: Organization ID
    """
    upload = await service.get_upload(upload_id, organization_id)
    if not upload:
        return

    processor = UploadProcessor(db, service)
    await processor.process_insertion(upload)
